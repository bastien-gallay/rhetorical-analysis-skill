"""Tests pour le générateur d'analyse rhétorique."""

import json
import tempfile
from pathlib import Path

import pytest

# Import relatif - ajuster selon la structure finale
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_analysis import (
    generate_xlsx,
    get_reliability_fill,
    create_main_analysis_sheet,
)
from openpyxl import Workbook


class TestReliabilityFill:
    """Tests pour la fonction de coloration selon fiabilité."""

    def test_high_reliability_green(self):
        fill = get_reliability_fill(5)
        assert fill.fgColor.rgb == "00C6EFCE"

    def test_medium_reliability_yellow(self):
        fill = get_reliability_fill(3)
        assert fill.fgColor.rgb == "00FFEB9C"

    def test_low_reliability_red(self):
        fill = get_reliability_fill(1)
        assert fill.fgColor.rgb == "00FFC7CE"

    def test_boundary_four_is_high(self):
        fill = get_reliability_fill(4)
        assert fill.fgColor.rgb == "00C6EFCE"

    def test_boundary_two_is_low(self):
        fill = get_reliability_fill(2)
        assert fill.fgColor.rgb == "00FFC7CE"


class TestGenerateXlsx:
    """Tests pour la génération complète du fichier XLSX."""

    @pytest.fixture
    def minimal_analysis(self):
        """Fixture avec une analyse minimale valide."""
        return {
            "metadata": {
                "title": "Test Article",
                "source": "https://example.com",
                "date_analysis": "2025-12-12"
            },
            "arguments": [
                {
                    "id": 1,
                    "label": "Test argument",
                    "original_text": "Lorem ipsum dolor sit amet.",
                    "claim": "Test claim",
                    "grounds": "Test grounds",
                    "warrant": "Test warrant",
                    "reasoning_type": "Deductive",
                    "fallacies": [],
                    "reliability": 4,
                    "reliability_rationale": "Well sourced",
                    "sources_cited": [],
                    "comment": ""
                }
            ],
            "synthesis": {
                "strengths": ["Good sourcing"],
                "weaknesses": ["Missing context"],
                "recurring_patterns": ["Appeal to authority"],
                "methodological_note": "Test note"
            }
        }

    @pytest.fixture
    def analysis_with_sources(self, minimal_analysis):
        """Fixture avec des sources CRAAP."""
        minimal_analysis["arguments"][0]["sources_cited"] = [
            {
                "name": "Test Source",
                "craap_score": {
                    "currency": 4,
                    "relevance": 5,
                    "authority": 4,
                    "accuracy": 3,
                    "purpose": 4
                }
            }
        ]
        return minimal_analysis

    def test_generate_xlsx_creates_file(self, minimal_analysis):
        """Vérifie que le fichier XLSX est créé."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "test.json"
            xlsx_path = Path(tmpdir) / "output.xlsx"
            
            json_path.write_text(json.dumps(minimal_analysis))
            
            result = generate_xlsx(str(json_path), str(xlsx_path))
            
            assert result is True
            assert xlsx_path.exists()

    def test_generate_xlsx_has_all_sheets(self, minimal_analysis):
        """Vérifie que toutes les feuilles sont créées."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "test.json"
            xlsx_path = Path(tmpdir) / "output.xlsx"
            
            json_path.write_text(json.dumps(minimal_analysis))
            generate_xlsx(str(json_path), str(xlsx_path))
            
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            
            expected_sheets = {
                "Analyse rhétorique",
                "Détail Toulmin", 
                "Évaluation sources (CRAAP)",
                "Synthèse",
                "Légende"
            }
            assert set(wb.sheetnames) == expected_sheets

    def test_generate_xlsx_invalid_json_path(self):
        """Vérifie le comportement avec un chemin JSON invalide."""
        result = generate_xlsx("/nonexistent/path.json", "/tmp/output.xlsx")
        assert result is False

    def test_generate_xlsx_invalid_json_content(self):
        """Vérifie le comportement avec un JSON malformé."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "invalid.json"
            xlsx_path = Path(tmpdir) / "output.xlsx"
            
            json_path.write_text("{ invalid json }")
            
            result = generate_xlsx(str(json_path), str(xlsx_path))
            assert result is False

    def test_sources_sheet_calculates_average(self, analysis_with_sources):
        """Vérifie le calcul de la moyenne CRAAP."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "test.json"
            xlsx_path = Path(tmpdir) / "output.xlsx"
            
            json_path.write_text(json.dumps(analysis_with_sources))
            generate_xlsx(str(json_path), str(xlsx_path))
            
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb["Évaluation sources (CRAAP)"]
            
            # La moyenne de [4, 5, 4, 3, 4] = 4.0
            avg_cell = ws.cell(row=2, column=8).value
            assert avg_cell == 4.0


class TestEmptyAnalysis:
    """Tests avec des analyses vides ou partielles."""

    def test_empty_arguments_list(self):
        """Vérifie le comportement avec une liste d'arguments vide."""
        data = {
            "metadata": {"title": "Empty"},
            "arguments": [],
            "synthesis": {}
        }
        
        with tempfile.TemporaryDirectory() as tmpdir:
            json_path = Path(tmpdir) / "empty.json"
            xlsx_path = Path(tmpdir) / "output.xlsx"
            
            json_path.write_text(json.dumps(data))
            result = generate_xlsx(str(json_path), str(xlsx_path))
            
            assert result is True
            assert xlsx_path.exists()
