#!/usr/bin/env python3
"""
Générateur de tableau d'analyse rhétorique

Ce script prend en entrée un fichier JSON contenant l'analyse structurée
et génère un fichier XLSX formaté.

Usage:
    python generate_analysis.py analysis.json output.xlsx
    
Le fichier JSON d'entrée doit avoir la structure suivante:
{
    "metadata": {
        "title": "Titre de l'article analysé",
        "source": "URL ou référence",
        "date_analysis": "2025-12-12",
        "analyst": "Claude"
    },
    "arguments": [
        {
            "id": 1,
            "label": "Résumé court de l'argument",
            "original_text": "Citation du texte original",
            "claim": "La thèse défendue",
            "grounds": "Les preuves avancées",
            "warrant": "Le lien logique implicite",
            "backing": "Support additionnel (optionnel)",
            "qualifier": "Nuances (optionnel)",
            "rebuttal": "Contre-arguments reconnus (optionnel)",
            "reasoning_type": "Type de raisonnement identifié",
            "fallacies": ["Liste des sophismes détectés"],
            "reliability": 4,
            "reliability_rationale": "Explication de la note",
            "sources_cited": [
                {
                    "name": "Nom de la source",
                    "craap_score": {
                        "currency": 4,
                        "relevance": 5,
                        "authority": 4,
                        "accuracy": 3,
                        "purpose": 4
                    }
                }
            ],
            "comment": "Commentaire critique (optionnel)"
        }
    ],
    "synthesis": {
        "strengths": ["Points forts du texte"],
        "weaknesses": ["Points faibles identifiés"],
        "recurring_patterns": ["Figures rhétoriques récurrentes"],
        "methodological_note": "Note sur le contexte du texte"
    }
}
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl non installé. Exécutez: pip install openpyxl")
    sys.exit(1)


# Styles
STYLES = {
    "header_font": Font(bold=True, color="FFFFFF", size=11),
    "header_fill": PatternFill("solid", fgColor="2F5496"),
    "wrap_align": Alignment(wrap_text=True, vertical="top"),
    "center_align": Alignment(wrap_text=True, vertical="top", horizontal="center"),
    "thin_border": Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ),
    "reliability_high": PatternFill("solid", fgColor="C6EFCE"),    # Vert (4-5)
    "reliability_medium": PatternFill("solid", fgColor="FFEB9C"),  # Jaune (3)
    "reliability_low": PatternFill("solid", fgColor="FFC7CE"),     # Rouge (1-2)
}


def get_reliability_fill(score: int) -> PatternFill:
    """Retourne le style de remplissage selon le score de fiabilité."""
    if score >= 4:
        return STYLES["reliability_high"]
    elif score == 3:
        return STYLES["reliability_medium"]
    else:
        return STYLES["reliability_low"]


def create_main_analysis_sheet(wb: Workbook, data: dict) -> None:
    """Crée la feuille principale d'analyse des arguments."""
    ws = wb.active
    ws.title = "Analyse rhétorique"
    
    # En-têtes
    headers = [
        "N°",
        "Argument traité",
        "Texte original (extrait)",
        "Thèse (Claim)",
        "Type de raisonnement",
        "Sophismes détectés",
        "Fiabilité (1-5)",
        "Évaluation de la fiabilité",
        "Commentaire"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = STYLES["header_font"]
        cell.fill = STYLES["header_fill"]
        cell.alignment = STYLES["center_align"]
        cell.border = STYLES["thin_border"]
    
    # Données
    for row_idx, arg in enumerate(data.get("arguments", []), 2):
        row_data = [
            arg.get("id", row_idx - 1),
            arg.get("label", ""),
            arg.get("original_text", "")[:500] + ("..." if len(arg.get("original_text", "")) > 500 else ""),
            arg.get("claim", ""),
            arg.get("reasoning_type", ""),
            "\n".join(arg.get("fallacies", [])) if arg.get("fallacies") else "Aucun détecté",
            arg.get("reliability", 3),
            arg.get("reliability_rationale", ""),
            arg.get("comment", "")
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = STYLES["wrap_align"]
            cell.border = STYLES["thin_border"]
            
            # Coloration fiabilité
            if col_idx == 7:
                cell.alignment = STYLES["center_align"]
                cell.fill = get_reliability_fill(value)
    
    # Largeurs de colonnes
    col_widths = [5, 25, 45, 40, 40, 25, 12, 45, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Hauteur des lignes
    for row in range(2, len(data.get("arguments", [])) + 2):
        ws.row_dimensions[row].height = 150


def create_toulmin_sheet(wb: Workbook, data: dict) -> None:
    """Crée une feuille détaillée avec la structure Toulmin complète."""
    ws = wb.create_sheet("Détail Toulmin")
    
    headers = ["N°", "Claim", "Grounds", "Warrant", "Backing", "Qualifier", "Rebuttal"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = STYLES["header_font"]
        cell.fill = STYLES["header_fill"]
        cell.alignment = STYLES["center_align"]
        cell.border = STYLES["thin_border"]
    
    for row_idx, arg in enumerate(data.get("arguments", []), 2):
        row_data = [
            arg.get("id", row_idx - 1),
            arg.get("claim", ""),
            arg.get("grounds", ""),
            arg.get("warrant", ""),
            arg.get("backing", "Non explicité"),
            arg.get("qualifier", "Non explicité"),
            arg.get("rebuttal", "Non reconnu")
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = STYLES["wrap_align"]
            cell.border = STYLES["thin_border"]
    
    col_widths = [5, 50, 50, 50, 40, 30, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_sources_sheet(wb: Workbook, data: dict) -> None:
    """Crée une feuille d'évaluation CRAAP des sources."""
    ws = wb.create_sheet("Évaluation sources (CRAAP)")
    
    headers = ["Source", "Arg. N°", "Currency", "Relevance", "Authority", "Accuracy", "Purpose", "Score moyen"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = STYLES["header_font"]
        cell.fill = STYLES["header_fill"]
        cell.alignment = STYLES["center_align"]
        cell.border = STYLES["thin_border"]
    
    row_idx = 2
    for arg in data.get("arguments", []):
        for source in arg.get("sources_cited", []):
            craap = source.get("craap_score", {})
            scores = [
                craap.get("currency", 0),
                craap.get("relevance", 0),
                craap.get("authority", 0),
                craap.get("accuracy", 0),
                craap.get("purpose", 0)
            ]
            avg_score = sum(scores) / len(scores) if scores else 0
            
            row_data = [
                source.get("name", ""),
                arg.get("id", ""),
                *scores,
                round(avg_score, 1)
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = STYLES["center_align"] if col_idx > 1 else STYLES["wrap_align"]
                cell.border = STYLES["thin_border"]
                
                # Coloration des scores
                if col_idx >= 3:
                    if isinstance(value, (int, float)):
                        cell.fill = get_reliability_fill(int(value))
            
            row_idx += 1
    
    col_widths = [40, 10, 12, 12, 12, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_synthesis_sheet(wb: Workbook, data: dict) -> None:
    """Crée la feuille de synthèse."""
    ws = wb.create_sheet("Synthèse")
    
    synthesis = data.get("synthesis", {})
    metadata = data.get("metadata", {})
    
    # Métadonnées
    ws['A1'] = "MÉTADONNÉES"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Titre: {metadata.get('title', 'Non spécifié')}"
    ws['A3'] = f"Source: {metadata.get('source', 'Non spécifié')}"
    ws['A4'] = f"Date d'analyse: {metadata.get('date_analysis', 'Non spécifié')}"
    
    # Synthèse
    ws['A6'] = "SYNTHÈSE CRITIQUE"
    ws['A6'].font = Font(bold=True, size=14)
    
    ws['A8'] = "Points forts:"
    ws['A8'].font = Font(bold=True)
    row = 9
    for strength in synthesis.get("strengths", []):
        ws[f'A{row}'] = f"• {strength}"
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Points faibles:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for weakness in synthesis.get("weaknesses", []):
        ws[f'A{row}'] = f"• {weakness}"
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Figures rhétoriques récurrentes:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for pattern in synthesis.get("recurring_patterns", []):
        ws[f'A{row}'] = f"• {pattern}"
        row += 1
    
    row += 2
    ws[f'A{row}'] = "Note méthodologique:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = synthesis.get("methodological_note", "")
    ws[f'A{row}'].alignment = STYLES["wrap_align"]
    
    ws.column_dimensions['A'].width = 100


def create_legend_sheet(wb: Workbook) -> None:
    """Crée la feuille de légende."""
    ws = wb.create_sheet("Légende")
    
    ws['A1'] = "ÉCHELLE DE FIABILITÉ"
    ws['A1'].font = Font(bold=True, size=14)
    
    legend = [
        (5, "Très haute", "Fait établi, consensus scientifique, sources multiples vérifiables"),
        (4, "Bonne", "Sources sérieuses, raisonnement logique valide, nuances possibles"),
        (3, "Moyenne", "Mélange faits/interprétations, sources partielles"),
        (2, "Faible", "Raisonnement contestable, sophismes identifiés"),
        (1, "Très faible", "Affirmations non sourcées, erreurs logiques majeures"),
    ]
    
    headers = ["Note", "Niveau", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = STYLES["header_font"]
        cell.fill = STYLES["header_fill"]
    
    for i, (note, niveau, desc) in enumerate(legend, 4):
        ws.cell(row=i, column=1, value=note).fill = get_reliability_fill(note)
        ws.cell(row=i, column=2, value=niveau)
        ws.cell(row=i, column=3, value=desc)
    
    ws['A10'] = "CRITÈRES CRAAP"
    ws['A10'].font = Font(bold=True, size=14)
    
    craap = [
        ("Currency", "L'information est-elle à jour ?"),
        ("Relevance", "L'information est-elle pertinente pour le propos ?"),
        ("Authority", "L'auteur/source est-il crédible dans ce domaine ?"),
        ("Accuracy", "Les faits sont-ils vérifiables et exacts ?"),
        ("Purpose", "Quelle est l'intention ? (informer, persuader, vendre...)"),
    ]
    
    for i, (critere, desc) in enumerate(craap, 12):
        ws.cell(row=i, column=1, value=critere).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60


def generate_xlsx(input_json: str, output_xlsx: str) -> bool:
    """
    Génère le fichier XLSX à partir du JSON d'analyse.
    
    Args:
        input_json: Chemin vers le fichier JSON d'entrée
        output_xlsx: Chemin vers le fichier XLSX de sortie
        
    Returns:
        True si succès, False sinon
    """
    try:
        with open(input_json, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"❌ Fichier non trouvé: {input_json}")
        return False
    except json.JSONDecodeError as e:
        print(f"❌ Erreur JSON: {e}")
        return False
    
    wb = Workbook()
    
    create_main_analysis_sheet(wb, data)
    create_toulmin_sheet(wb, data)
    create_sources_sheet(wb, data)
    create_synthesis_sheet(wb, data)
    create_legend_sheet(wb)
    
    try:
        wb.save(output_xlsx)
        print(f"✅ Fichier généré: {output_xlsx}")
        return True
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde: {e}")
        return False


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_analysis.py <input.json> <output.xlsx>")
        print("\nExemple:")
        print("  python generate_analysis.py analysis.json rapport_analyse.xlsx")
        sys.exit(1)
    
    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]
    
    success = generate_xlsx(input_json, output_xlsx)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
